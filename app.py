import openpyxl as xl
from openpyxl.chart import BarChart,Reference

def work_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet=wb['Sheet1']


    for row in range(2,sheet.max_row+1):
        cell=sheet.cell(row,3)
        correcte_price=cell.value*0.9
        correcte_price_cell=sheet.cell(row,4)
        correcte_price_cell.value=correcte_price

    values=Reference(sheet,min_row=2,
            max_row=sheet.max_row,
            min_col=4,
            max_col=4)
    chart=BarChart()
    chart.add_data(values)
    sheet.add_chart(chart,'e2')

    wb.save(filename)

